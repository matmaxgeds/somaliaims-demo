from django.core.management.base import BaseCommand
from management.models import Sector
from openpyxl import load_workbook



class Command(BaseCommand):
    help = 'Imports sector information from xlsx. The xlsx file is given as a cmd line argument'
    SECTOR_RANGE = 'B8:C44'

    def add_arguments(self, parser):
        parser.add_argument('file_path')

    def handle(self, file_path, *args, **options):
        wb = load_workbook(file_path, use_iterators=True)
        sheet = wb.get_sheet_by_name('Sectors')
        for row in sheet.iter_rows(self.SECTOR_RANGE):
            count = 0
            for cell in row:
                if count == 0 or 0 == (count % 2):
                    try:
                        full_name = cell.value
                        values = full_name.rpartition(': ')
                        code = values[0]
                        name = values[2].strip()
                        print(name)
                    except TypeError:  # Raised by running re.search on an empty value
                        code = ''
                count += 1
                if 0 != (count % 2):
                    description = cell.value.rpartition(': ')[2]
                try:
                    Sector.objects.get(name=name)
                except Exception:
                    if name:
                        new_sector = Sector(name=name, description=description)
                        new_sector.save()

