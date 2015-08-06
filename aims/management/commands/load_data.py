from django.core.management.base import BaseCommand
from management.models import Sector, Organization, Currency
from data_entry.models import Project, Document, SectorAllocation, LocationAllocation, SubLocation
from openpyxl import load_workbook
from django.contrib.auth.models import User


class Command(BaseCommand):
    help = 'Imports sector and project information'
    PROJECT_RANGE = 'B2:AZ537'
    SECTOR_RANGE = 'B8:C44'

    def add_arguments(self, parser):
        parser.add_argument('file_path')

    def handle(self, file_path, *args, **options):
        wb = load_workbook(file_path, use_iterators=True)
        sheet = wb.get_sheet_by_name('Sectors')
        for row in sheet.iter_rows(self.SECTOR_RANGE):
            count = 0
            for cell in row:
                if count == 0 or 0 == (count % 2):
                    try:
                        full_name = cell.value
                        values = full_name.rpartition(': ')
                        code = values[0]
                        name = values[2].strip()
                        print(name)
                    except TypeError:  # Raised by running re.search on an empty value
                        code = ''
                count += 1
                if 0 != (count % 2):
                    description = cell.value.rpartition(': ')[2]
                try:
                    Sector.objects.get(name=name)
                except Exception:
                    if name:
                        new_sector = Sector(name=name, description=description)
                        new_sector.save()

        sheet = wb.get_sheet_by_name('2015 Projects')
        tuple(sheet.iter_rows('B2:AZ537'))
        for row in tuple(sheet.iter_rows('B2:AZ537')):
            for cell in row:
                print(cell.value)
                break

        # print(str(sheet.rows))
        # for row in sheet.iter_rows(self.PROJECT_RANGE):
        #     user = User.objects.get(id=1)
        #     count = 0
        #     max_count = 23  # Count of last column to be fetched.
        #     for cell in row:
        #         if count > max_count:
        #             count = 0
        #         if count == 0:
        #             reporter = cell.value
        #             print(reporter)
        #             count += 1
        #             continue
        #         fund = []
        #         if count == 1:
        #             funders = cell.value
        #             s = funders.rpartition(': ')
        #             n = s[2].split(', ')
        #             for org in n:
        #                 try:
        #                     o = Organization.objects.get(short_name=org, name=org)
        #                     fund.append(o)
        #                 except Exception:
        #                     new_save = Organization(name=org, short_name=org)
        #                     print(str(new_save))
        #                     new_save.save()
        #                     fund.append(new_save)
        #             print(funders)
        #             count += 1
        #             continue
        #         plementers = []
        #         if count == 2:
        #             implementers = cell.value
        #             d = implementers.rpartition(': ')
        #             l = d[2].split(', ')
        #             for org in l:
        #                 try:
        #                     o = Organization.objects.get(short_name=org, name=org)
        #                     plementers.append(o)
        #                 except Exception:
        #                     new_save = Organization(name=org, short_name=org)
        #                     new_save.save()
        #                     plementers.append(new_save)
        #             print(implementers)
        #             count += 1
        #             continue
        #         if count == 3:
        #             title = cell.value
        #             print(title)
        #             count += 1
        #             continue
        #         if count == 4:
        #             description = cell.value
        #             print(description)
        #             count += 1
        #             continue
        #         if count == 5:
        #             ocha_ots = cell.value
        #             print(ocha_ots)
        #             count += 1
        #             continue
        #         if count == 6:
        #             sector = cell.value
        #             count += 1
        #             continue
        #         if count == 7:
        #             sector1 = cell.value
        #             print(sector1)
        #             if type(sector1) is not str:
        #                 pass
        #             else:
        #                 sec1 = Sector(name=sector1)
        #                 sec1.save()
        #             count += 1
        #             continue
        #         if count == 8:
        #             sector1_allocation = cell.value
        #             # alloc1 = SectorAllocation()
        #             print(sector1_allocation)
        #             count += 1
        #             continue
        #         if count == 9:
        #             sector2 = cell.value
        #             print(sector2)
        #             if type(sector2) is not str:
        #                 pass
        #             else:
        #                 sec2 = Sector(name=sector2)
        #                 sec2.save()
        #             count += 1
        #             continue
        #         if count == 10:
        #             sector2_allocation = cell.value
        #             # alloc2 = SectorAllocation()
        #             # print(sector2_allocation)
        #             count += 1
        #             continue
        #         if count == 11:
        #             startDate = cell.value
        #             print(startDate)
        #             #print(str(datetime.datetime.strptime(startDate, "%d%m%Y").date()))
        #             count += 1
        #             continue
        #         if count == 12:
        #             endDate = cell.value
        #             print(endDate)
        #             #print(str(datetime.datetime.strptime(endDate, "%d%m%Y").date()))
        #             count += 1
        #             continue
        #         if count == 13:  # Reporting currency.
        #             currency = cell.value
        #             try:
        #                 s = Currency.objects.get(currency=currency)
        #                 c = s
        #             except Exception:
        #                 d = Currency(abbrev=currency, currency=currency)
        #                 d.save()
        #                 c = d
        #             print(c)
        #             count += 1
        #             continue
        #         if count == 14:  # Reporting currency rate per USD.
        #             rateToUSD = cell.value
        #             print(rateToUSD)
        #             count += 1
        #             continue
        #         if count == 15:
        #             value = cell.value
        #             print(value)
        #             count += 1
        #             continue
        #         if count == 16:
        #             spendingLastYear = cell.value
        #             #project['spendingLastYear'] = spendingLastYear
        #             print(spendingLastYear)
        #             count += 1
        #             continue
        #         if count == 17:
        #             spendingToDate = cell.value
        #             #project['spendingToDate'] = spendingToDate
        #             print(spendingToDate)
        #             count += 1
        #             continue
        #         # if count == 13:
        #         #     spendingThisYear = cell.value
        #         #     print(spendingThisYear)
        #         #     count += 1
        #         #     continue
        #         if count == 18:
        #             spendingNextYear = cell.value
        #             #project['spendingNextYear'] = spendingNextYear
        #             print(spendingNextYear)
        #             count += 1
        #             continue
        #         if count == 19:
        #             locationSpending = cell.value
        #             localloc = LocationAllocation()
        #             print(locationSpending)
        #             count += 1
        #             continue
        #         if count == 20:
        #             sublocations = cell.value
        #             sub = SubLocation()
        #             print(sublocations)
        #             count += 1
        #             continue
        #         if count == 21:
        #             documentLinks = cell.value
        #             doc = Document(link_to_document_website=documentLinks)
        #             print(documentLinks)
        #             count += 1
        #             continue
        #         if count == 22:
        #             sdrf_ssa = cell.value
        #             print(sdrf_ssa)
        #             count += 1
        #             continue
        #
        #     project = Project(user=user, name=title, description=description,
        #                       ocha_fts_reported=True if ocha_ots == 'Yes' else False, startDate=startDate, endDate=endDate, currency=c, rateToUSD=rateToUSD, value=value, sdrf_ssa=sdrf_ssa)
        #     project.save()
        #     for i in fund:
        #         project.funders.add(i)
        #     for j in plementers:
        #         project.implementers.add(j)

                # if count == 23:
                #     contactName = cell.value
                #     print(contactName)
                #     count += 1
                #     continue
                # if count == 19:
                #     contactEmail = cell.value
                #     print(contactEmail)
                #     count += 1
                #     continue
                # if count == 20:
                #     contactPhone = cell.value
                #     print(contactPhone)
                #     count += 1
