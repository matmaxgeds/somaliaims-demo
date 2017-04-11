__author__ = 'alphabuddha'

from openpyxl import load_workbook
import sys

RANGE = 'F4:BD446'

# Test if user includes path to file as argument
try:
    xlsfile = sys.argv[1]
except IndexError:
    sys.stderr.write("The correct way to run the script is python importer.py path/to/xls.xlsx \n")

    exit(0)


# Get workbook and worksheets
wb = load_workbook(xlsfile, read_only=True)
worksheets = wb.get_sheet_names()


def getSectors():
    # if conn:
    #    cur = conn.cursor()  # Cursor for DB operations
    wb = load_workbook('/home/alphabuddha/Downloads/xls.xlsx', read_only=True)
    sheet = wb.get_sheet_by_name('Sectors')
    for row in sheet.iter_rows('B8:C44'):
        count = 0
        for cell in row:
            if count == 0 or 0 == (count % 2):
                try:
                    full_name = cell.value
                    values = full_name.rpartition(':')
                    code = values[0]
                    name = values[2].strip()
                    print(code)
                    print(name)
                except TypeError:  # Raised by running re.search on an empty value
                    code = ''
            if 0 != (count % 2):
                description = cell.value
                print(description)
            count += 1


def get_projects():
    # if conn:
    #    cur = conn.cursor()  # Cursor for DB operations
    sheet = wb.get_active_sheet()
    for row in sheet.iter_rows(RANGE):
        count = 0
        max_count = 20  # Count of last column to be fetched.
        for cell in row:
            if count > max_count:
                count = 0
            if count == 0:
                reporter = cell.value
                print(reporter)
                count += 1
                continue
            if count == 1:
                funders = cell.value
                print(funders)
                count += 1
                continue
            if count == 2:
                implementers = cell.value
                print(implementers)
                count += 1
                continue
            if count == 3:
                title = cell.value
                print(title)
                count += 1
                continue
            if count == 4:
                description = cell.value
                print(description)
                count += 1
                continue
            if count == 5:
                ocha_ots = cell.value
                print(ocha_ots)
                count += 1
                continue
            if count == 6:
                sectors = cell.value
                print(sectors)
                count += 1
                continue
            if count == 7:
                startDate = cell.value
                print(startDate)
                count += 1
                continue
            if count == 8:
                endDate = cell.value
                print(endDate)
                count += 1
                continue
            if count == 9:
                currency = cell.value
                print(currency)
                count += 1
                continue
            if count == 10:
                rateToUSD = cell.value
                print(rateToUSD)
                count += 1
                continue
            if count == 11:
                value = cell.value
                print(value)
                count += 1
                continue
            if count == 12:
                spendingLastYear = cell.value
                print(spendingLastYear)
                count += 1
                continue
            if count == 13:
                spendingThisYear = cell.value
                print(spendingThisYear)
                count += 1
                continue
            if count == 14:
                spendingNextYear = cell.value
                print(spendingNextYear)
                count += 1
                continue
            if count == 15:
                locationSpending = cell.value
                print(locationSpending)
                count += 1
                continue
            if count == 16:
                sublocations = cell.value
                print(sublocations)
                count += 1
                continue
            if count == 17:
                documentLinks = cell.value
                print(documentLinks)
                count += 1
                continue
            if count == 18:
                contactName = cell.value
                print(contactName)
                count += 1
                continue
            if count == 19:
                contactEmail = cell.value
                print(contactEmail)
                count += 1
                continue
            if count == 20:
                contactPhone = cell.value
                print(contactPhone)
                count += 1
                continue


if __name__ == '__main__':
    getSectors()
    get_projects()

# Manual db connection using psycopg2

# import psycopg2
# import logging

# DB connection
# try:
#     conn = psycopg2.connect("dbname='somaliaims' user='somaliaims' host='localhost' password='aims.somali'")
# except Exception:
#     conn = psycopg2.connect(
#        "host='somaliaims-43.postgres.pythonanywhere-services.com' password='aims.somali' port='10043' dbname='somaliaims'  user='somaliaims'")
# except:
#     logfile = 'db.log'
#     logging.basicConfig(filename=logfile, level=logging.DEBUG)
     # Make it send email to maintainer
#     sys.stderr.write("An error occured trying to connect to the database \n")
#     logging.debug('Error connecting to db')
#     exit(0)
